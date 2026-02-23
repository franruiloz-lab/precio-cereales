"""
Scraper: Junta de Castilla y León - Excel semanal de lonjas.

Fuente: https://agriculturaganaderia.jcyl.es/web/es/estadistica-informacion-agraria/lonjas-2026.html
Frecuencia: Semanal (lunes), con ~2 semanas de retraso.
Formato: XLSX con hoja "Granos y henos".

Lonjas incluidas: Ebro, Albacete, Ciudad Real, Lérida, Salamanca,
                  Valladolid, Palencia, Lerma, Segovia, León.
Precios en: €/100Kg (se convierten a €/t multiplicando x10).
"""

import os
import re
import tempfile
import requests
import openpyxl
from bs4 import BeautifulSoup


INDEX_URL_TEMPLATE = (
    "https://agriculturaganaderia.jcyl.es/web/es/"
    "estadistica-informacion-agraria/lonjas-{anio}.html"
)

# Nombres de cereales tal como aparecen en el Excel (fila izquierda)
CEREALES_EXCEL = [
    "Trigo",
    "Cebada (+64)",
    "Cebada (60-64)",
    "Maiz seco nac",
    "Avena",
    "Centeno",
    "Trigo Duro",
    "Pipa de girasol",
]


def _buscar_url_excel(anio, semana):
    """Busca la URL del Excel de una semana concreta en la página índice."""
    url = INDEX_URL_TEMPLATE.format(anio=anio)
    resp = requests.get(url, timeout=30, verify=False)
    resp.raise_for_status()

    soup = BeautifulSoup(resp.text, "html.parser")

    # Los enlaces a los Excel tienen texto como "Semana 7" o "semana 7"
    patron = re.compile(rf"semana\s*{semana}\b", re.IGNORECASE)
    for enlace in soup.find_all("a", href=True):
        texto = enlace.get_text(strip=True)
        if patron.search(texto):
            href = enlace["href"]
            if not href.startswith("http"):
                href = "https://agriculturaganaderia.jcyl.es" + href
            return href

    # Intento alternativo: buscar por URL que contenga el número de semana
    for enlace in soup.find_all("a", href=True):
        href = enlace["href"]
        if f"semana%20{semana}" in href.lower() or f"semana+{semana}" in href.lower():
            if not href.startswith("http"):
                href = "https://agriculturaganaderia.jcyl.es" + href
            return href

    return None


def _descargar_excel(url):
    """Descarga el Excel a un archivo temporal y devuelve la ruta."""
    resp = requests.get(url, timeout=60, verify=False)
    resp.raise_for_status()

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(resp.content)
    tmp.close()
    return tmp.name


def _parsear_granos(filepath):
    """
    Parsea la hoja 'Granos y henos' del Excel de la JCyL.

    Retorna: dict de {nombre_lonja: {nombre_cereal: precio_eur_tonelada}}
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # Buscar la hoja de granos
    hoja = None
    for nombre in wb.sheetnames:
        if "grano" in nombre.lower() or "heno" in nombre.lower():
            hoja = wb[nombre]
            break

    if hoja is None:
        raise ValueError(f"No se encontró hoja de granos. Hojas: {wb.sheetnames}")

    # Leer todas las celdas en una matriz
    filas = []
    for row in hoja.iter_rows(values_only=True):
        filas.append(list(row))

    # Encontrar la fila de nombres de lonjas (contiene "Ebro", "Albacete", etc.)
    fila_lonjas = None
    idx_lonjas = -1
    for i, fila in enumerate(filas):
        textos = [str(c).strip() if c else "" for c in fila]
        texto_unido = " ".join(textos)
        if "Ebro" in texto_unido and ("Salamanca" in texto_unido or "Le" in texto_unido):
            fila_lonjas = fila
            idx_lonjas = i
            break

    if fila_lonjas is None:
        raise ValueError("No se encontró la fila de nombres de lonjas")

    # Extraer posiciones de columna de cada lonja
    # Cada lonja ocupa 2 columnas: precio + variación
    lonjas_cols = {}
    for j, celda in enumerate(fila_lonjas):
        if celda and str(celda).strip():
            nombre = str(celda).strip()
            # Limpiar nombre
            nombre = nombre.replace("Lonja de ", "").replace("Lonja del ", "")
            nombre = nombre.replace("\n", " ").strip()
            if nombre and nombre not in ("DOCUMENTACIÓN PÚBLICA", "Lonjas nacionales",
                                          "Lonjas regionales", ""):
                lonjas_cols[nombre] = j

    # Buscar filas de cereales (empiezan después de la fila de lonjas + fecha)
    resultado = {nombre: {} for nombre in lonjas_cols}

    for i in range(idx_lonjas + 2, len(filas)):
        fila = filas[i]
        if not fila or not fila[0]:
            # Revisar columna 1 también
            nombre_producto = str(fila[1]).strip() if len(fila) > 1 and fila[1] else ""
            if not nombre_producto:
                continue
        else:
            nombre_producto = str(fila[0]).strip()
            if not nombre_producto:
                continue

        # Buscar si coincide con algún cereal conocido
        cereal_match = None
        for cereal in CEREALES_EXCEL:
            if cereal.lower() in nombre_producto.lower():
                cereal_match = cereal
                break

        # Coincidencia más específica
        if nombre_producto.strip().startswith("Trigo") and "Duro" in nombre_producto:
            cereal_match = "Trigo Duro"
        elif nombre_producto.strip().startswith("Trigo") and "Duro" not in nombre_producto:
            cereal_match = "Trigo"

        if cereal_match is None:
            # Si encontramos la nota de pie, paramos
            if "diferencia" in nombre_producto.lower() or "nota" in nombre_producto.lower():
                break
            continue

        # Extraer precios de cada lonja
        for nombre_lonja, col_idx in lonjas_cols.items():
            try:
                valor = fila[col_idx]
                if valor is not None and valor != "" and str(valor).strip().upper() != "S/C":
                    precio_100kg = float(valor)
                    precio_tonelada = round(precio_100kg * 10, 2)
                    resultado[nombre_lonja][cereal_match] = precio_tonelada
            except (ValueError, TypeError, IndexError):
                continue

    # Limpiar lonjas sin datos
    resultado = {k: v for k, v in resultado.items() if v}

    return resultado


def obtener_precios_jcyl(anio, semana):
    """
    Obtiene precios de cereales del Excel semanal de la Junta de CyL.

    Returns:
        dict: {nombre_lonja: {nombre_cereal: precio_eur_tonelada}} o None si no hay datos.
    """
    print(f"    Buscando Excel semana {semana}/{anio}...")
    url = _buscar_url_excel(anio, semana)

    if url is None:
        print(f"    No se encontró archivo para semana {semana}")
        return None

    print(f"    Descargando: {url}")
    filepath = _descargar_excel(url)

    try:
        precios = _parsear_granos(filepath)
        return precios
    finally:
        os.unlink(filepath)
